import openpyxl

if __name__ == '__main__':
    wb = openpyxl.load_workbook('streets.xlsx', data_only=True)
    sh = wb.worksheets[0]

    row = 1
    while True:
        pos = sh.cell(row=row, column=1).value
        if not pos: break
        data = sh.cell(row=row, column=2).value
        num, typeopbj, first, *rest = data.split()
        if len(rest) == 3:
            first += ' ' + rest[0]
            rest = rest[1:]
        print(pos, typeopbj, first, ' '.join(rest))
        sh.cell(row=row, column=3).value = int(num.replace('.', ''))
        sh.cell(row=row, column=4).value = typeopbj
        sh.cell(row=row, column=5).value = first
        sh.cell(row=row, column=6).value = ' '.join(rest)
        row += 1
    wb.save('streets.xlsx')
    wb.close()
