import openpyxl

if __name__ == '__main__':
    wb = openpyxl.load_workbook('streets_old.xlsx', data_only=True)
    sh = wb.worksheets[0]

    row = 1
    while True:
        num = sh.cell(row=row, column=1).value
        if not num: break
        pos = 'м. Шепетівка'
        olds = sh.cell(row=row, column=2).value
        old = olds.split()
        if old[-1] in ('вулиця', 'провулок', 'площа'):
            objtype = old[-1]
            old_name = ' '.join(old[:-1])
        else:
            objtype = ""
            old_name = olds
        rename_date = sh.cell(row=row, column=3).value
        news = sh.cell(row=row, column=4).value
        if news:
            new = news.split()
            if new[-1] in ('вулиця', 'провулок', 'площа'):
                new_name = ' '.join(new[:-1])
            else:
                new_name = news
        else:
            new_name = ''
        applied = 1
        tag = 'old'
        
        sh.cell(row=row, column=6).value = num
        sh.cell(row=row, column=7).value = pos
        sh.cell(row=row, column=8).value = objtype
        sh.cell(row=row, column=9).value = old_name
        sh.cell(row=row, column=10).value = new_name
        sh.cell(row=row, column=11).value = rename_date
        sh.cell(row=row, column=12).value = applied
        sh.cell(row=row, column=13).value = tag

        row += 1
    wb.save('streets_old.xlsx')
    wb.close()
